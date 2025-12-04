import pandas as pd
import requests
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

KEYWORDS = [
    "404", "not found", "page not found", "error", "missing",
    "このページは存在しません", "見つかりません",
    "ページが見つかりません", "エラー",
    "không tồn tại", "không tìm thấy", "lỗi", "trang không tồn tại"
]

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)",
    "Mozilla/5.0 (X11; Linux x86_64)"
]

icons = {
    "load": "📥",
    "process": "⚙️",
    "request": "🌐",
    "export": "📤",
    "done": "✅",
    "error": "❌"
}

def log(stage, text):
    print(f"{icons.get(stage, '•')} {text}")

def is_soft_404(content, status):
    text = content.lower()
    if status == 404:
        return True
    return any(k in text for k in KEYWORDS)

def fetch(url):
    headers = {
        "User-Agent": random.choice(USER_AGENTS),
        "Cache-Control": "no-cache",
        "Pragma": "no-cache"
    }
    return requests.get(url, timeout=10, headers=headers)

def check_url(url):
    try:
        log("request", f"Đang yêu cầu: {url}")
        res = fetch(url)

        if is_soft_404(res.text, res.status_code):
            return url, 404, "Trang không tồn tại (Soft-404)"

        return url, int(res.status_code), "Tồn tại (OK)"

    except Exception as e:
        return url, None, f"Lỗi kết nối: {e}"

def generate_excel(results, output_path):
    df = pd.DataFrame(results, columns=["URL", "Trạng thái", "Ghi chú"])
    df.to_excel(output_path, index=False)
    format_excel(output_path)

def format_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 50

    center = Alignment(horizontal="center", vertical="center")
    white = Font(color="FFFFFF")

    green = PatternFill(start_color="92D050", fill_type="solid")
    red = PatternFill(start_color="FF0000", fill_type="solid")
    yellow = PatternFill(start_color="FFFF00", fill_type="solid")

    ws.row_dimensions[1].height = 25

    for row in ws.iter_rows(min_row=2):
        url_cell, status_cell, note_cell = row

        ws.row_dimensions[url_cell.row].height = 30

        url_cell.alignment = Alignment(vertical="center")
        status_cell.alignment = center

        status = status_cell.value

        if status == 200:
            status_cell.fill = green

        elif status == 404:
            status_cell.fill = red
            status_cell.font = white

        else:
            status_cell.fill = yellow

    wb.save(path)
